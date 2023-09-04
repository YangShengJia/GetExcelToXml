import openpyxl
import xml.etree.ElementTree as ET
import re
from langdetect import detect

# Read data from Excel column#/Users/07607.ben.yang/HidinString/androidneedhindi.xlsx
droid_excel_file_path = "/Users/07607.ben.yang/HidinString/androidneedhindi.xlsx"
ios_excel_file_path = "/Users/07607.ben.yang/HidinString/iOSMissStringExcel.xlsx"
excel_column_index = 0  # Index of the column you want to compare
excel_column_index_value_engb = 1  # Index of the column you want to compare
droid_excel_data_keys = []
droid_excel_data_values = []
ios_excel_data_keys = []
ios_excel_data_values = []

droidworkbook = openpyxl.load_workbook(droid_excel_file_path)
droidworksheet = droidworkbook.active

iosworkbook = openpyxl.load_workbook(ios_excel_file_path)
iosworksheet = iosworkbook.active

for row in droidworksheet.iter_rows(min_row=1, max_row=droidworksheet.max_row,
                                   min_col=excel_column_index + 1,
                                    max_col=excel_column_index + 1):
    droid_excel_data_keys.append(row[0].value)

for row in droidworksheet.iter_rows(min_row=1, max_row=droidworksheet.max_row,
                                   min_col=excel_column_index_value_engb + 1,
                                    max_col=excel_column_index_value_engb + 1):

    droid_excel_data_values.append(row[0].value)
    
for row in iosworksheet.iter_rows(min_row=1, max_row=iosworksheet.max_row,
                               min_col=excel_column_index + 1,
                               max_col=excel_column_index + 1):
    ios_excel_data_keys.append(row[0].value)    

for row in iosworksheet.iter_rows(min_row=1, max_row=iosworksheet.max_row,
                               min_col=excel_column_index_value_engb + 1,
                                 max_col=excel_column_index_value_engb + 1):

    ios_excel_data_values.append(row[0].value)
    
print(f"droidexceldatakeys: {droid_excel_data_keys}")
print(f"droidexceldatavalue: {droid_excel_data_values}")
print(f"iosexceldatakeys: {ios_excel_data_keys}")
print(f"iosexceldatavalue: {ios_excel_data_values}")


for droid_value in droid_excel_data_values:
    
    if droid_value in ios_excel_data_values:
        matching_key_value = ios_excel_data_keys[droid_excel_data_values.index(droid_value)]
        matching_key_value = ios_excel_data_keys[ios_excel_data_values.index(droid_value)]
        
        print(f"Matching name: {matching_key_value}")
        '''
        change_value = excel_data_value[excel_data_item.index(name)]

        detected_language = detect(text)
        if detected_language == "en":
            string_element.text = change_value
        '''
#       print(f"Matching name: {name}")
#       print(f"Matching text: {text}")
