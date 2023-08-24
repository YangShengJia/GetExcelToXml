import openpyxl
import xml.etree.ElementTree as ET
import re
from langdetect import detect

def detect_special_characer(pass_string): 
    regex= re.compile('[@_!#$%^&*()<>?/\|}{~:]') 
    if(regex.search(pass_string) == None): 
        res = False
    else: 
        res = True
    return(res)
    
def safe_startswith(text, prefix):
    if text is None:
        return False
    return text.startswith(prefix)

# Read data from Excel column
excel_file_path = "/Users/07607.ben.yang/testexcel.xlsx"
excel_column_index = 2  # Index of the column you want to compare
excel_column_index_value = 3  # Index of the column you want to compare
excel_column_index_value_engb = 4  # Index of the column you want to compare en-gb
excel_data_item = []
excel_data_value = []
excel_data_engb_value = []
workbook = openpyxl.load_workbook(excel_file_path)
worksheet = workbook.active
for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row,
                               min_col=excel_column_index + 1, max_col=excel_column_index + 1):
    excel_data_item.append(row[0].value)

for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row,
                               min_col=excel_column_index_value + 1,
                               max_col=excel_column_index_value + 1):
    excel_data_value.append(row[0].value)
    
for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row,
                               min_col=excel_column_index_value_engb + 1,
                               max_col=excel_column_index_value_engb+ 1):
    excel_data_engb_value.append(row[0].value)
    
print(f"exceldata: {excel_data_item}")
print(f"exceldatavalue: {excel_data_value}")
print(f"exceldataengbvalue: {excel_data_engb_value}")
# Parse XML file
xml_file_path = "/Users/07607.ben.yang/HidinString/strings.xml"
tree = ET.parse(xml_file_path)
root = tree.getroot()

# Compare XML data with Excel data
for string_element in root.findall("string"):
    name = string_element.get("name")
    text = string_element.text#get xml value
    if name in excel_data_item:
        matching_value = excel_data_item[excel_data_item.index(name)]       
        change_value = excel_data_value[excel_data_item.index(name)]
        if	change_value != "#N/A":
            string_element.text = change_value
        else:
            na_value = change_value
#       print(f"Matching name: {name}")
#       print(f"Matching text: {text}")

for string_element in root.findall("string"):
    text = string_element.text#get xml value
    if text in excel_data_engb_value:
        matching_value = excel_data_engb_value[excel_data_engb_value.index(text)]
        #print(f"Matching name: {matching_value}")
        change_value = excel_data_value[excel_data_engb_value.index(text)]
        #print(f"Matching Hindi name:{matching_value} : {change_value}")
        if change_value != "#N/A":
            string_element.text = change_value
        else :
            na_value = change_value

#tree.write(xml_file_path)
#tree.write(xml_file_path, encoding="utf-8", xml_declaration=True)
#detect language

for string_element in root.findall("string"):
    name = string_element.get("name")#get xml name
    text = string_element.text#get xml value
    
    if  text == None:
        noneText = None
        print(f"Eng {name}; {noneText}")
    elif safe_startswith(text, "+"):
        plusString = text
        print(f"Eng SYMBOL PLUS {name}; {plusString}")
    elif safe_startswith(text, "0"):
        numberString = text
        print(f"Eng NUMBER {name}; {numberString}")
    elif detect_special_characer(text):
        specialStirng = text
        print(f"Eng special character {name}; {specialStirng}")        
    else:
        detected_language = detect(text)
        if detected_language == "en":
            print(f"Eng ELSE {name}; {text}")
            
print("Comparison done.")

